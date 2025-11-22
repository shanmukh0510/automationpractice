import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook[sheetname]
    return worksheet.max_row
def getColumnCoUNt(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook[sheetname]
    return worksheet.max_colimn
def ReadDataFromExcel(file,sheetname,rowno,columnno):
    workbook =  openpyxl.load_workbook(file)
    worksheet = workbook[sheetname]
    return worksheet.cell(rowno,columnno).value
def WriteDataToExcel(file,sheetname,rowno,columnno,data):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook[sheetname]
    worksheet.cell(rowno,columnno).value = data
    workbook.save(file)




