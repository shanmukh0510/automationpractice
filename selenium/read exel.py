import openpyxl

file_path = "D:\Financial Sample.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook["Sheet1"]

rows = sheet.max_row
columns = sheet.max_column
print(rows)
print(columns)

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r,c).value, end= "        ")
    print(       )



